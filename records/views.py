from django.shortcuts import render,redirect, get_object_or_404
from .models import Record, Detail, Holiday
from django.http import HttpResponse
import openpyxl
import datetime
import time

# Create your views here.
def import_view(request):
	if "GET" == request.method:
		return render(request, 'records/import_file.html', {})
	else:
		excel_file = request.FILES["excel_file"]
		
		wb = openpyxl.load_workbook(excel_file)
		
		worksheet = wb["Sheet1"]
		print(worksheet)
		
		sheet = wb.active
		print(sheet)
		i=2
		
		user_data = list()

		while(i<sheet.max_row):
			#new_obj1 = Record()
			user_dict = {}
			att_data = list()
			user_dict['s_no'] = sheet.cell(i,column=2).value
			user_dict['s_name'] = sheet.cell(i,column=8).value

			#new_obj1.staff_num = sheet.cell(i,column=2).value
			#new_obj1.staff_name = sheet.cell(i,column=8).value
			#new_obj1.save()
			j=i+2
			for k in range(1,32,1):
				attendance = {}
				#new_obj2 = Detail()
				attendance['attendance_date'] = sheet.cell(j,column=1).value
				attendance['attendance_day'] = sheet.cell(j,column=2).value
				attendance['attendance_entry'] = sheet.cell(j,column=5).value
				attendance['attendance_exit'] = sheet.cell(j,column=6).value
				
				#new_obj2.record = new_obj1

				att_data.append(attendance['attendance_date'].strftime('%Y/%m/%d'))
				att_data.append(attendance['attendance_day'])

				#new_obj2.work_date = attendance['attendance_date'].strftime('%Y/%m/%d')
				#new_obj2.work_day = attendance['attendance_day']
				
				if attendance['attendance_entry']!=None :
					att_data.append(attendance['attendance_entry'].strftime('%H:%M'))
					#new_obj2.entry_time = attendance['attendance_entry'].strftime('%H:%M')
				else:
					att_data.append(attendance['attendance_entry'])
					#new_obj2.entry_time = attendance['attendance_entry']
				
				if attendance['attendance_exit']!=None :
					att_data.append(attendance['attendance_exit'].strftime('%H:%M'))
					#new_obj2.exit_time = attendance['attendance_exit'].strftime('%H:%M')
				else:
					att_data.append(attendance['attendance_exit'])
					#new_obj2.exit_time = attendance['attendance_exit']

				#new_obj2.save()
				j=j+1
				user_dict['total_attendance'] = att_data
			user_data.append(user_dict)
			
			i=i+36
		context={
			'user_data':user_data
		}
		return render(request, 'records/import_file.html', context)

def all_names_view(request):
	obj = Record.objects.all()
	context ={
	'my_obj':obj,
	}

	return render(request,'records/list_names.html',context)

def detail_view(request,my_id):
	record_obj = Record.objects.get(id=my_id)
	detail_obj = Detail.objects.filter(record=record_obj)
	holiday_obj = Holiday.objects.all()
	context = {
	'obj':detail_obj,
	'my_obj':record_obj,
	'h_obj':holiday_obj,
	}

	return render(request,'records/detail.html',context)

def holiday_view(request):
	htypes = request.POST.getlist('htype')
	hdates = request.POST.getlist('hdate')
	for i in range(0,len(htypes)):
		holiday_object = Holiday()
		holiday_object.holiday_type = htypes[i]
		p = datetime.datetime.strptime(hdates[i],'%Y-%m-%d')		
		holiday_object.date=p.strftime('%Y/%m/%d')
		holiday_object.save()
	return redirect('/import/')

def update_view(request,this_id):
	this_id = str(this_id)
	detail_id = request.POST['det_id']
	detail_obj = Detail.objects.get(id=detail_id)
	k = request.POST['work']
	if k=="P":
		detail_obj.entry_time = request.POST['in_time']
		detail_obj.exit_time = request.POST['out_time']
		detail_obj.save()
	elif k=="L":
		detail_obj.remark = request.POST['remark']
		detail_obj.reason = k
		detail_obj.save()
	elif k=="OD":
		detail_obj.remark = request.POST['remark']
		detail_obj.reason = k
		detail_obj.save()
	elif k=="WFH":
		detail_obj.remark = request.POST['remark']
		detail_obj.reason = k
		detail_obj.save()
	return redirect('/list/detail.html/'+this_id)